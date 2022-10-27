from selenium.webdriver.common.by import By
from selenium import webdriver
import time
import pandas as pd
import openpyxl
import xlsxwriter
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementClickInterceptedException


class TrendyolTest:

    def __init__(self):
        self.browser = webdriver.Chrome()
        self.repository = []
        self.allExcelInfo = []
        self.allReport = []
        self.allReportState = []

    def getExcelInfo(self):
        path = "trendyolTest.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        for row in range(0, sheet_obj.max_row):
            for col in sheet_obj.iter_cols(1, sheet_obj.max_column):
                
                trendyolTest.allExcelInfo.append(col[row].value)

    def openPage(self):
        url = "https://www.trendyol.com/"
        self.browser.get(url)
        time.sleep(2)
        self.browser.find_element(
            By.XPATH, "//button[@id='onetrust-accept-btn-handler']").click()
        print("ana sayfa acildi")
        trendyolTest.allReport.append("1-ana sayfa acildi")
        trendyolTest.allReportState.append("OK-1")
        time.sleep(2)

    def addBasket(self):
        controlException = True
        for i in range(len(trendyolTest.allExcelInfo)):
            time.sleep(1)
            self.browser.find_element(
                By.XPATH, "//input[@class='vQI670rJ']").click()
            self.browser.find_element(
                By.XPATH, "//input[@class='vQI670rJ']").clear()
            print("arama butonuna  girildi")
            trendyolTest.allReport.append("2-arama butonuna  girildi")
            trendyolTest.allReportState.append("OK-2")
            thewords = self.browser.find_element(
                By.XPATH, "//input[@class='vQI670rJ']")
            time.sleep(2)

            thewords.send_keys(trendyolTest.allExcelInfo[i])
            print("3-obje yazildi")
            trendyolTest.allReport.append("3-obje yazildi")
            trendyolTest.allReportState.append("OK-3")
            time.sleep(1)

            self.browser.find_element(
                By.XPATH, "//i[@class='ft51BU2r']").click()
            print("arama butonu calisti")
            trendyolTest.allReport.append("4-arama butonu calisti")
            trendyolTest.allReportState.append("OK-4")
            time.sleep(3)

            try:

                self.browser.find_element(
                    By.XPATH, "//div[text() = 'Kargo Bedava' ]").click()

                print("kargo bedava")
                trendyolTest.allReport.append("5-kargo bedava")
                trendyolTest.allReportState.append("OK-5")
            except NoSuchElementException:
                print("NoSuchElementException Hatasi")
                trendyolTest.allReportState.append("OK-5 -NSEE")
                controlException = False

            except:
                print("Bilinmeyen bir hata meydana geldi")
                trendyolTest.allReportState.append("OK-5 -NotKnowException")
                self.browser.refresh()
                time.sleep(1)
                
                self.browser.find_element(
                    By.XPATH, "//div[text() = 'Kargo Bedava' ]").click()

            # hatali olmasi ve bitmis ise donguden cik ,eger hata var ama daha fazla urun varsa devam et.
            if (controlException == False):
                self.browser.execute_script("window.history.go(-1)")
                time.sleep(1)
                break
            time.sleep(1)
            allspans = []
            spansnumber = []
            time.sleep(2)

            for span in self.browser.find_elements(By.XPATH, "//span[@class='ratingCount']"):
                time.sleep(0.1)
                span1 = span.text.replace("(", "")
                span2 = span1.replace(")", "")
                allspans.append(int(span2))
                spansnumber.append(span)

            time.sleep(1)

            print(allspans)
            time.sleep(3)

            maxItem = max(allspans)
            indexItem = allspans.index(maxItem)
            print(indexItem)
            time.sleep(2)

            newstatment = "//span[text() = '{}' ]"
            state = newstatment.format(maxItem)
            print("item secildi.")
            trendyolTest.allReport.append("6-item secildi.")
            trendyolTest.allReportState.append("OK-6")
            self.browser.find_element(By.XPATH, state).click()
            time.sleep(1)
            self.browser.switch_to.window(self.browser.window_handles[1])
            time.sleep(1)

            self.browser.find_element(
                By.CLASS_NAME, "add-to-basket").click()
            print("item sepete eklendi.")
            trendyolTest.allReport.append("7-item sepete eklendi")
            trendyolTest.allReportState.append("OK-7")
            time.sleep(2)
            self.browser.close()
            self.browser.switch_to.window(self.browser.window_handles[0])
            time.sleep(1)

    def showBasket(self):
        try:
            time.sleep(1)
            self.browser.find_element(
                By.XPATH, "//p[text() ='Sepetim']").click()
            print("sepete girildi")
        except ElementClickInterceptedException:
            self.browser.refresh()
            time.sleep(2)
            self.browser.find_element(
                By.XPATH, "//p[text() ='Sepetim']").click()
            print("sepete girildi")

    def getReport(self):
        workbook = xlsxwriter.Workbook('write_list.xlsx')
        worksheet = workbook.add_worksheet()
        print(trendyolTest.allReport)
        for i in range(len(trendyolTest.allExcelInfo)):
            worksheet.write(0, i+1, trendyolTest.allExcelInfo[i])
        for j in range(7):
            worksheet.write(j+1, 0, trendyolTest.allReport[j])

        for t in range(len(trendyolTest.allExcelInfo)):
            for z in range(7):
                worksheet.write(z+1, t+1, trendyolTest.allReportState[z])

        workbook.close()


trendyolTest = TrendyolTest()

trendyolTest.getExcelInfo()
trendyolTest.openPage()
trendyolTest.addBasket()
trendyolTest.showBasket()
trendyolTest.getReport()
